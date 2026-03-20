"""CSV and Excel content extractor — row-range chunking."""

import csv
import io

from openpyxl import load_workbook

from src.core.extractors.base import Chunk, ExtractionResult

DEFAULT_CHUNK_ROWS = 100


class CsvExcelExtractor:
    """Extract normalized text chunks from CSV or Excel content."""

    def extract(self, raw: bytes, config: dict | None = None) -> ExtractionResult:
        """Extract text from CSV or Excel, chunked by row ranges.

        Config keys:
            content_type: str — "csv" or "xlsx" (required)
            chunk_rows: int — rows per chunk (default: 100)
            sort_keys: list[str] — column names to sort by before chunking
        """
        config = config or {}
        content_type = config.get("content_type", "csv")
        chunk_size = config.get("chunk_rows", DEFAULT_CHUNK_ROWS)
        sort_keys = config.get("sort_keys")

        if content_type == "xlsx":
            header, rows = self._parse_xlsx(raw)
        else:
            header, rows = self._parse_csv(raw)

        if sort_keys:
            rows = self._sort_rows(rows, header, sort_keys)

        if not rows:
            return ExtractionResult(chunks=[])

        return self._chunk_rows(header, rows, chunk_size)

    def _parse_csv(self, raw: bytes) -> tuple[list[str], list[list[str]]]:
        """Parse CSV bytes into header and data rows."""
        text = raw.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        rows_iter = iter(reader)
        try:
            header = next(rows_iter)
        except StopIteration:
            return [], []
        rows = [row for row in rows_iter if any(cell.strip() for cell in row)]
        return header, rows

    def _parse_xlsx(self, raw: bytes) -> tuple[list[str], list[list[str]]]:
        """Parse Excel bytes into header and data rows."""
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        try:
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = [str(c) if c is not None else "" for c in next(rows_iter)]
            except StopIteration:
                return [], []
            rows = [
                [str(c) if c is not None else "" for c in row]
                for row in rows_iter
                if any(c is not None for c in row)
            ]
            return header, rows
        finally:
            wb.close()

    def _sort_rows(
        self, rows: list[list[str]], header: list[str], sort_keys: list[str]
    ) -> list[list[str]]:
        """Sort rows by specified column names."""
        indices = []
        for key in sort_keys:
            if key in header:
                indices.append(header.index(key))
        if not indices:
            return rows
        return sorted(
            rows,
            key=lambda row: tuple(row[i] if i < len(row) else "" for i in indices),
        )

    def _chunk_rows(
        self, header: list[str], rows: list[list[str]], chunk_size: int
    ) -> ExtractionResult:
        """Split rows into chunks of chunk_size."""
        chunks = []

        for i in range(0, len(rows), chunk_size):
            batch = rows[i : i + chunk_size]
            start = i + 1
            end = i + len(batch)
            text = self._rows_to_csv_text(header, batch)

            chunks.append(
                Chunk(
                    index=len(chunks),
                    chunk_type="row_range",
                    label=f"Rows {start}-{end}",
                    text=text,
                )
            )

        return ExtractionResult(chunks=chunks)

    def _rows_to_csv_text(self, header: list[str], rows: list[list[str]]) -> str:
        """Serialize header + rows to properly quoted CSV text."""
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(header)
        writer.writerows(rows)
        return buf.getvalue().strip()
